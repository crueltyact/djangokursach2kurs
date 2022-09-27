"""
Модуль парсинга матрицы компетенций
"""
import datetime
from typing import Union

import openpyxl
import xlrd


def get_parents(matrix: list, r: int) -> list[str, str, str]:
    """
    Функция получения компетенций

    :param matrix: матрица данных
    :type matrix: list
    :param r: строка компетенции
    :type r: int
    :return: список данных о компетенции
    :rtype: list[3]
    """
    scd = matrix[r][1].replace('\n', '')
    fst = matrix[r][0].replace('\n', '')
    return [fst, scd, matrix[r][2]]


def get_info_for_table(matrix: list, rng: list, c: int) -> list:
    """
    Функция получения информации для составления таблицы компетенций

    :param matrix: матрица данных
    :type matrix: list
    :param rng: список строк для каждой компетенции
    :type rng: list
    :param c: строка названия компетенции
    :type c: int
    :return: список данных для построения таблицы компетенций
    :rtype: list
    """
    res = [
        {
            'competency_code': '',
            'competency_name': '',
            'indicators': [['', set()]]
        }
    ]
    for r in rng:
        if matrix[r][c] == '+':
            f_code, s_code, t_code = get_parents(matrix, r)
            if f_code == '' or s_code == '' or t_code == '':
                continue
            code, name = [el.strip() for el in list(filter(bool, f_code.split('.')))]
            if res[-1]['competency_code'] != code:
                res.append({
                    'competency_code': code,
                    'competency_name': name,
                    'indicators': [['', set()]]
                })
                res[-1]['indicators'][0][0] = s_code
                res[-1]['indicators'][0][1].add(t_code)
            else:
                if res[-1]['indicators'][-1][0] != s_code:
                    res[-1]['indicators'].append(['', set()])
                    res[-1]['indicators'][-1][0] = s_code
                    res[-1]['indicators'][-1][1].add(t_code)
                else:
                    res[-1]['indicators'][-1][1].add(t_code)
    del res[0]
    return res


def get_ranges(matrix: list) -> list:
    """
    Функция получения строк, относящихся к одной компетенции

    :param matrix: матрица данных
    :type matrix: list
    :return: список номеров строк для компетенции
    :rtype: list
    """
    rows = len(matrix)
    skill_types = []
    k = 0
    for i in range(rows)[1::]:
        if matrix[i][2] == matrix[i][1] == matrix[i][0] and matrix[i][0] != '':
            skill_types += [range(k, i)]
            k = i
    del skill_types[0]
    skill_types += [range(k, rows)]
    return skill_types


def parse_title(txt: str) -> dict[str, Union[str, int]]:
    """
    Функция парсинга заголовка файла

    :param txt: строка заголовка файла
    :type txt: str
    :return: список данных, полученных из заголовка
    :rtype: dict[str, Union[str, int]]
    """
    import re
    res = {}
    txt = re.sub('[»«]', '"', txt)
    txt = re.sub('[\n,]', '', txt)
    mas = txt.split('"')
    res['profile_name'] = mas[5]
    res['program_code'] = mas[3]
    txt = re.sub('["]', ' ', txt)
    mas = txt.split()
    for el in mas:
        if el.count('.') == 2:
            res['program_code'] = el + ' ' + res['program_code']
        if el.count('/') == 1:
            temp = el.split('/')
            res['year_start'] = temp[0]
            res['year_end'] = temp[1]
    return res


def get_matrix(filename: str) -> list:
    """
    Функция получения матрицы из файла

    :param filename: путь к файлу
    :type filename: str
    :return: список данных из файла
    :rtype: list
    """
    xls = xlrd.open_workbook(filename)
    xls = xls.sheet_by_index(0)
    mx_row, mx_column = xls.nrows, xls.ncols
    wb = openpyxl.load_workbook(filename)
    sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    all_data = []
    for row_index in range(1, mx_row + 1):
        row = []
        for col_index in range(1, mx_column + 1):
            vals = sheet.cell(row_index, col_index).value
            if vals is None:
                for crange in sheet.merged_cells:
                    clo, rlo, chi, rhi = crange.bounds
                    top_value = sheet.cell(rlo, clo).value
                    if rlo <= row_index <= rhi and clo <= col_index <= chi:
                        vals = top_value
                        break
            row.append(vals)
        if len(list(filter(bool, row))) > 0:
            all_data.append(row)
    for i in range(len(all_data)):
        for j in range(len(all_data[0])):
            if all_data[i][j] is None:
                all_data[i][j] = ''
            all_data[i][j] = str(all_data[i][j]).strip()
    return all_data


def get_data_from_matrix(data: dict, title: dict, matrix: list, skill_types: list, c: int, key: str) -> dict:
    """
    Функция заполнения списка данных по шаблону

    :param data: данные, получаемые из матрицы
    :type data: dict
    :param title: данные из заголовка файла
    :type title: dict
    :param matrix: данные из матрицы данных файла
    :type matrix: list
    :param skill_types: список строк компетенции
    :type skill_types: list
    :param c: текущий столбец парсинка
    :type c: int
    :param key: название учебной дисциплины
    :type key: str
    :return: оюновлённые данные из матрицы
    :rtype: dict
    """
    data[key] = {}
    data[key]['program_name'] = key
    data[key]['profile_name'] = title['profile_name']
    data[key]['program_code'] = title['program_code']
    data[key]['year_start'] = title['year_start']
    data[key]['year_end'] = title['year_end']
    data[key]['current_year'] = str(datetime.date.today().year)
    data[key]['part_type'] = str.lower(matrix[1][c])
    universal_competences = get_info_for_table(matrix, skill_types[0], c)
    general_professional_competencies = get_info_for_table(matrix, skill_types[1], c)
    professional_competencies = get_info_for_table(matrix, skill_types[2], c)
    if len(universal_competences) > 0:
        data[key]['universal_competences'] = universal_competences
    if len(general_professional_competencies) > 0:
        data[key]['general_professional_competencies'] = general_professional_competencies
    if len(professional_competencies) > 0:
        data[key]['professional_competencies'] = professional_competencies
    return data


def get_info_from_excel(filename: str) -> tuple[dict, list]:
    """
    Функция получения информации из матрицы компетенций

    :param filename: путь к файлу
    :type filename: str
    :return: список полученных данных и список названий учебных дисциплин
    :rtype: tuple[dict, list]
    """
    matrix = get_matrix(filename)
    skill_types = get_ranges(matrix)
    for i in range(len(matrix))[::-1]:
        if len(list(filter(bool, matrix[i]))) == 0:
            del matrix[i]
    cols = len(matrix[0])
    title = parse_title(matrix[0][0])
    data = {}
    key_data = []
    for c in range(cols)[3::]:
        key = matrix[2][c]
        if key == '':
            continue
        key_data.append(key)
        data = get_data_from_matrix(data, title, matrix, skill_types, c, key)
    return data, key_data
